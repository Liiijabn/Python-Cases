from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
import datetime

# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
#
# ws['B1'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")


wb = load_workbook('你的好朋友是谁v1.2.xlsx')
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']
processed_sheet = wb['Sheet']
friends_num = 'friends_num'                # 好友数
be_friends_num = 'be_friends_num'          # 粉丝数
homo_friends_num = 'homo_friends_num'      # 同性好友数
hetero_friends_num = 'hetero_friends_num'  # 异性好友数
homo_be_friends_num = 'homo_be_friends_num'        # 同性粉丝数
hetero_be_friends_num = 'hetero_be_friends_num'    # 异性粉丝数

all_child = {}

for child in sheet2['C2':'E42']:
    all_child[child[0].value] = {}
    all_child[child[0].value]['gender'] = child[1].value
    all_child[child[0].value]['nick_name'] = child[2].value
    all_child[child[0].value]['friends'] = set()
    all_child[child[0].value]['be_friend'] = set()
    all_child[child[0].value]['friend_each_other'] = set()
    all_child[child[0].value][friends_num] = 0            # 朋友数
    all_child[child[0].value][homo_friends_num] = 0       # 同性朋友数
    all_child[child[0].value][hetero_friends_num] = 0     # 异性朋友数
    all_child[child[0].value][be_friends_num] = 0         # 粉丝数
    all_child[child[0].value][homo_be_friends_num] = 0    # 同性粉丝数
    all_child[child[0].value][hetero_be_friends_num] = 0  # 异性粉丝数


for child in sheet1['C2':'P42']:
    name = child[0].value
    friends = set([c.value for c in child[1:] if c.value is not None])
    all_child[name]['friends'] = friends
    for l in friends:
        all_child[l]['be_friend'].add(name)

for child, relation in all_child.items():
    all_child[child]['friend_each_other'] = all_child[child]['friends'] & all_child[child]['be_friend']
    for f in relation['friends']:
        all_child[child][friends_num] += 1
        if all_child[f]['gender'] == all_child[child]['gender']:
            all_child[child][homo_friends_num] += 1
        else:
            all_child[child][hetero_friends_num] += 1

    for f in relation['be_friend']:
        all_child[child][be_friends_num] += 1
        if all_child[f]['gender'] == all_child[child]['gender']:
            all_child[child][homo_be_friends_num] += 1
        else:
            all_child[child][hetero_be_friends_num] += 1


processed_sheet['A2'] = '姓名'
processed_sheet['B2'] = '小名'
processed_sheet['C2'] = '性别'
processed_sheet['D2'] = '好友数'
processed_sheet['E2'] = '同性好友数'
processed_sheet['F2'] = '异性好友数'

processed_sheet['G2'] = '粉丝数'
processed_sheet['H2'] = '同性粉丝数'
processed_sheet['I2'] = '异性粉丝数'

processed_sheet['J2'] = '互粉数'

processed_sheet['K2'] = '好友'
processed_sheet['L2'] = '粉丝'
processed_sheet['M2'] = '互粉'


row_number = 2
for child, relation in all_child.items():
    print(child, relation)
    row_number += 1
    processed_sheet['A%d' % row_number].value = child
    processed_sheet['B%d' % row_number].value = relation['nick_name']
    processed_sheet['C%d' % row_number].value = relation['gender']
    processed_sheet['D%d' % row_number].value = relation[friends_num]
    processed_sheet['E%d' % row_number].value = relation[homo_friends_num]
    processed_sheet['F%d' % row_number].value = relation[hetero_friends_num]
    processed_sheet['G%d' % row_number].value = relation[be_friends_num]
    processed_sheet['H%d' % row_number].value = relation[homo_be_friends_num]
    processed_sheet['I%d' % row_number].value = relation[hetero_be_friends_num]
    processed_sheet['J%d' % row_number].value = len(relation['friend_each_other'])
    friends = set([all_child[friend]['nick_name'] for friend in relation['friends'] if friend is not None])
    processed_sheet['K%d' % row_number].value = ','.join(friends)
    be_friends = set([all_child[be_friend]['nick_name'] for be_friend in relation['be_friend'] if be_friend is not None])
    processed_sheet['L%d' % row_number].value = ','.join(be_friends)
    fs = set([all_child[f]['nick_name'] for f in relation['friend_each_other'] if f is not None])
    processed_sheet['M%d' % row_number].value = ','.join(fs)


wb.save('result.xlsx')

# print(all_child)


# for i in childs:
#     print('%s - %s ' % (i[0].value, i[1].value))

# relations = wb['Sheet1']['C2':'P42']
#
# row_num = 1
# for row in relations:
#     row_num += 1
#     child = row[0]
#     friends = row[1:]
#     friends_num = 0
#     homosexual_friends_num = 0
#     heterosexual_friends_num = 0
#     child_gender = genders[child.value]
#
#     for f in friends:
#         if f.value is None:
#             break
#         else:
#             friends_num += 1
#             # print(f.color)
#             if genders[f.value] == child_gender:
#                 homosexual_friends_num += 1
#             else:
#                 heterosexual_friends_num += 1
#
#     print(child.value, ' friends： %s 同性朋友数：%d 异性朋友数：%d' % (friends_num, homosexual_friends_num, heterosexual_friends_num))
#
# print(row_num)
#
#
#
